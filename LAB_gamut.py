#  此段说明，LAB 色域边界不是u'v'色域边界, 通过python 3.12.3测试 ， 2024-12-19
import numpy as np
import colour
import pandas as pd
import matplotlib.pyplot as plt
from scipy.spatial import ConvexHull
from openpyxl import Workbook
from colour.plotting import plot_chromaticity_diagram_CIE1976UCS
from gamut_util import ConvexHull_general, ConvexHull_uni , LCHab_to_XYZ, LCHuv_to_XYZ, XYZ_to_uv, write_excel
import matplotlib

matplotlib.use('TkAgg')
plt.ion()
print("Currently Using Backend:", matplotlib.get_backend()) 

ILLUM_C = colour.CCS_ILLUMINANTS["CIE 1931 2 Degree Standard Observer"]['C']  # CIE xy chromaticity coordinates
D65 = colour.CCS_ILLUMINANTS["CIE 1931 2 Degree Standard Observer"]["D65"]    # CIE xy chromaticity coordinates
D50 = colour.CCS_ILLUMINANTS["CIE 1931 2 Degree Standard Observer"]["D50"]    # CIE xy chromaticity coordinates

#from colour.adaptation import chromatic_adaptation
from colour.adaptation import chromatic_adaptation_VonKries

# 转换为XYZ白点 (假设 Y = 1)
def xy_to_XYZ(xy, Y=1.0):
    x, y = xy
    X = (Y / y) * x
    Z = (Y / y) * (1 - x - y)
    return np.array([X, Y, Z])

D50_white = xy_to_XYZ(D50)  # D50白点的XYZ值
D65_white = xy_to_XYZ(D65)  # D65白点的XYZ值
ILLUM_C_white = xy_to_XYZ(ILLUM_C)

from openpyxl import load_workbook
def delete_sheet_if_exists(file_path, sheet_name):
    # 使用 openpyxl 加载 Excel 文件
    workbook = load_workbook(file_path)
    # 如果目标工作表已经存在，则删除
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    # 保存删除后的文件
    workbook.save(file_path)

import colorspacious as cs
# 定义 CIELAB 色彩的映射函数
def hue_to_lab_color(hue_angle):
    # CIELAB 颜色空间的 a 和 b 轴可以用来表示色相
    L = 70  # 这里选择固定的亮度 L*
    C = 50  # 固定的色度 C*
    # 将 hue_angle 转换为弧度制并计算 a* 和 b* 值
    a = C * np.cos(np.radians(hue_angle))
    b = C * np.sin(np.radians(hue_angle))
    # CIELAB 颜色空间下的 (L*, a*, b*) 转换为 RGB
    lab_color = cs.cspace_convert([L, a, b], "CIELab", "sRGB1")
    # 确保 RGB 值在 [0, 1] 范围内
    lab_color = np.clip(lab_color, 0, 1)
    return lab_color

ILLUM_S = D50  #   确定标准光源  两坐标:  ILLUM_C,  D50,  D65,  ******************************************************
ILLUM_S_white = D50_white  #  三坐标:    ILLUM_C_white,  D50_white,  D65_white
file_path = 'LCH_gamut.xlsx'  #  Excel file   *****************************
sheet_name = '12640_LCHab'                           #  12640_LCHab,  TC300_LCHab,  pointer_LCHab, pointer_LCHuv    HP_printer,  PhotoGamut   *******************
source_format = 'LCHab'                            # LCHab,  LCHuv
# Read Excel data
df = pd.read_excel(file_path, sheet_name=sheet_name, index_col=0)
# Extract hue angles, lightness values, chroma values
hue = df.index.to_numpy()  # Hue angles (0-350, first column)
L_values = df.columns.to_numpy(dtype=float)  # Lightness values (5, 10, ..., 95, first row)
Chroma = df.to_numpy()  # Chroma values (36x19 array)

# hue: (36,) 数组，表示色相角
# L_values: (19,) 数组，表示亮度值
# Chroma: (36, 19) 数组，表示色相角与亮度对应的色度值
# 创建一个列表来存储结果
LCH_list = []
hue_angle_list = []
# 遍历每个色相角和亮度值的组合
for i, L in enumerate(L_values):  # 遍历每个亮度值
    for j, h in enumerate(hue):  # 遍历每个色相角
        C = Chroma[j, i]  # 对应的色度值 (Chroma value at (hue, L))
        # 创建 L, C, h 的数组
        LCH = np.array([L, C, h])
        # 将 LCH 数组添加到列表中
        LCH_list.append(LCH)
        hue_angle_list.append(h)
# 转换为 NumPy 数组，形状为 (36*19, 3)
LCH_array = np.array(LCH_list)
hue_angle_array = np.array(hue_angle_list)

# 投影到 L = 0 的 ab 平面
# 计算 a 和 b 的值
a_values = LCH_array[:, 1] * np.cos(np.radians(LCH_array[:, 2]))  # C * cos(h)
b_values = LCH_array[:, 1] * np.sin(np.radians(LCH_array[:, 2]))  # C * sin(h)

# 创建 ab 平面数据
ab_points = np.vstack((a_values, b_values)).T

# 使用 ConvexHull 找到边界
hull = ConvexHull(ab_points)

# 绘制所有点
plt.figure(figsize=(8, 6))
plt.scatter(a_values, b_values, color='blue', s=10, label='All Points')

# 绘制边界点
for simplex in hull.simplices:
    plt.plot(ab_points[simplex, 0], ab_points[simplex, 1], 'r-')
plt.scatter(ab_points[hull.vertices, 0], ab_points[hull.vertices, 1], color='red', label='Boundary Points')

# 绘制 CIE 1976UCS 色域图
plot_chromaticity_diagram_CIE1976UCS(show_diagram_colours=False, color ='lightgray', standalone=False) 
# 强制保持坐标轴比例为1:1
plt.gca().set_aspect('equal', adjustable='box')
plt.axis((-0.1, 0.7, -0.1, 0.7))
# 转换Lab到u'v'，绘制convex_hull在u'v'平面边界点。
Lab_points = np.vstack((LCH_array[:, 0], a_values, b_values)).T
XYZ = colour.Lab_to_XYZ(Lab_points, ILLUM_S)  
u_prime, v_prime = XYZ_to_uv(XYZ)
u_prime_hull = u_prime[hull.vertices]
v_prime_hull = v_prime[hull.vertices]
plt.scatter(u_prime, v_prime , color='lightgray')  # 所有颜色点
plt.scatter(u_prime_hull, v_prime_hull , color='red' )  # Lab 边界点

# 设置标签
#plt.xlabel("a*")
#plt.ylabel("b*")
#plt.title("Color Points and Convex Hull in a*b* Plane")
plt.legend()
plt.grid()
plt.axis('equal')
plt.show()
aa = 1
