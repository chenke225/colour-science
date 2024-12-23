import numpy as np
import pandas as pd
import matplotlib
import matplotlib.pyplot as plt
from scipy.spatial import ConvexHull
from openpyxl import Workbook
from colour.plotting import plot_chromaticity_diagram_CIE1976UCS
import colour
# 色域绘制，标注色度角， convex hull 原始算法找外轮廓。 uv坐标转换用colour-science函数， 绘图添加hue角度标注。将产生的外轮廓写入excel。
# 通过python3.12.3测试，2024-12-19

matplotlib.use('TkAgg')
plt.ion()
print("Currently Using Backend:", matplotlib.get_backend())  # 当前默认后端

# Placeholder values for D50 white point
ILLUM_C = colour.CCS_ILLUMINANTS["CIE 1931 2 Degree Standard Observer"]['C']
D65 = colour.CCS_ILLUMINANTS["CIE 1931 2 Degree Standard Observer"]["D65"]
D50 = colour.CCS_ILLUMINANTS["CIE 1931 2 Degree Standard Observer"]["D50"]
ILLUM_S = D50  #   确定标准光源  ******************************************************
file_path = 'LCH_gamut.xlsx'  # Replace with the path to your Excel file   *****************************
sheet_name = '12640_LCHab'                                   # ************************************************************************
Data_format = 'LCHab'

def XYZ_to_u_v(XYZ):
    
    # 确保 XYZ 至少是二维数组，形状变为 (n, 3)
    XYZ = np.atleast_2d(XYZ)# 提取 X, Y, Z 列
    X = XYZ[:,0]
    Y = XYZ[:,1]
    Z = XYZ[:,2]
    # 计算分母：X + 15Y + 3Z
    denominator = X + 15 * Y + 3 * Z
    # 避免除以零的情况，添加一个非常小的值 epsilon
    epsilon = 1e-10
    denominator = np.maximum(denominator, epsilon)
    # 计算 u' 和 v' 按公式
    u_prime = 4 * X / denominator
    v_prime = 9 * Y / denominator
    return u_prime, v_prime

def LCHab_to_uv_1(LCHab, ILLUMINATNAT):
    Lab = colour.LCHab_to_Lab(LCHab)
    XYZ = colour.Lab_to_XYZ(Lab, ILLUMINATNAT)  # ILLUM_C
    return XYZ_to_u_v(XYZ)

def LCHab_to_uv_2(LCHab, ILLUMINATNAT):
    Lab = colour.LCHab_to_Lab(LCHab)
    XYZ = colour.Lab_to_XYZ(Lab, ILLUMINATNAT)  # ILLUM_C   
    Luv = colour.XYZ_to_Luv(XYZ, ILLUMINATNAT)
    uv = colour.Luv_to_uv(Luv)
    return uv

def LCHuv_to_uv(LCHuv, ILLUMINATNAT):
    Luv = colour.LCHuv_to_Luv(LCHuv)
    XYZ = colour.Luv_to_XYZ(Luv, ILLUMINATNAT)  # ILLUM_C
    xy_3d = colour.XYZ_to_xyY(XYZ)  # 第一种转换
    xy_2d = xy_3d[:2]
    return colour.xy_to_Luv_uv(xy_2d)

# Read Excel data
df = pd.read_excel(file_path, sheet_name=sheet_name, index_col=0)

# Extract hue angles, lightness values, chroma values
hue = df.index.to_numpy()  # Hue angles (0-350, first column)
L_values = df.columns.to_numpy(dtype=float)  # Lightness values (5, 10, ..., 95, first row)
Chroma = df.to_numpy()  # Chroma values (36x19 array)

#  试验比较LCHab 到 u'v'转换 #############################################################
# Convert all LAB points to u'v'
# 假设 LCHab_to_uv_1 和 LCHab_to_uv_2 已定义

def compare_uv_results(L_values, hue, Chroma, ILLUM_S):
    u_prime_list_1 = []
    v_prime_list_1 = []
    u_prime_list_2 = []
    v_prime_list_2 = []
    hue_angle_list = []
    
    # 遍历所有 L 和 hue 值，计算 uv_prime 值
    for i, L in enumerate(L_values):
        for j, h in enumerate(hue):
            C = Chroma[j, i]  # 取色度值 (h_ab, L)
            LCH = np.array([L, C, h])
            
            # 使用 LCHab_to_uv_1 进行转换
            uv_prime_1 = LCHab_to_uv_1(LCH, ILLUM_S)
            u_prime_1, v_prime_1 = uv_prime_1[0], uv_prime_1[1]
            u_prime_list_1.append(u_prime_1)
            v_prime_list_1.append(v_prime_1)
            
            # 使用 LCHab_to_uv_2 进行转换
            uv_prime_2 = LCHab_to_uv_2(LCH, ILLUM_S)
            u_prime_2, v_prime_2 = uv_prime_2[0], uv_prime_2[1]
            u_prime_list_2.append(u_prime_2)
            v_prime_list_2.append(v_prime_2)
            
            # 存储 hue 角度
            hue_angle_list.append(h)
    
    # 将列表转换为 numpy 数组以便于比较
    u_prime_array_1 = np.array(u_prime_list_1)
    v_prime_array_1 = np.array(v_prime_list_1)
    u_prime_array_2 = np.array(u_prime_list_2)
    v_prime_array_2 = np.array(v_prime_list_2)
    
    # 使用 np.allclose 来比较 u' 和 v' 是否在容差范围内相等
    u_equal = np.allclose(u_prime_array_1, u_prime_array_2, atol=1e-6)
    v_equal = np.allclose(v_prime_array_1, v_prime_array_2, atol=1e-6)
    
    # 输出比较结果
    if u_equal and v_equal:
        print("LCHab_to_uv_1 和 LCHab_to_uv_2 转换结果一致。")
    else:
        print("LCHab_to_uv_1 和 LCHab_to_uv_2 转换结果不一致。")


# 调用函数比较转换结果
compare_uv_results(L_values, hue, Chroma, ILLUM_S)


# 1. No Convex Hull: Find the max C*ab for each hue angle and calculate u'v'
u_prime_no_hull = []
v_prime_no_hull = []
hue_angle_no_hull = []
for j, h in enumerate(hue):
    # Find the maximum C*ab for each hue angle across all lightness values
    max_C_ab = np.max(Chroma[j, :])
    L_max_index = np.argmax(Chroma[j, :])
    L_max = L_values[L_max_index]  # Get corresponding L* value
    LCH = np.array([L_max, max_C_ab, h])
    if Data_format == 'LCHab':
        uv_prime = LCHab_to_uv_1(LCH, ILLUM_S)  # uv_prime = LCHab_to_uv(LCHab, ILLUM_S)
    else:
        uv_prime = LCHuv_to_uv(LCH, ILLUM_S)
    u_prime, v_prime = uv_prime[0], uv_prime[1]

    # Store results
    u_prime_no_hull.append(u_prime)
    v_prime_no_hull.append(v_prime)
    hue_angle_no_hull.append(h)
# Now ensure the curve is closed by duplicating the 0-degree point at 360 degrees
u_prime_no_hull.append(u_prime_no_hull[0])
v_prime_no_hull.append(v_prime_no_hull[0])
hue_angle_no_hull.append(360)


# 2  Convert all LAB points to u'v' for Convex Hull
u_prime_list = []
v_prime_list = []
hue_angle_list = []

# Convert all LAB points to u'v'
for i, L in enumerate(L_values):
    for j, h in enumerate(hue):
        C = Chroma[j, i]  # Chroma value at (h_ab, L)
        LCH = np.array([L, C, h])
        if Data_format == 'LCHab':
            uv_prime = LCHab_to_uv_1(LCH, ILLUM_S)  # uv_prime = LCHab_to_uv(LCHab, ILLUM_S)
        else:
            uv_prime = LCHuv_to_uv(LCH, ILLUM_S)
        u_prime, v_prime = uv_prime[0], uv_prime[1]   ###############
        u_prime_list.append(u_prime)
        v_prime_list.append(v_prime)
        hue_angle_list.append(h)

# Convert to numpy array for convex hull computation
uv_points = np.column_stack((u_prime_list, v_prime_list))

# 使用ConvexHull函数
hull = ConvexHull(uv_points)
# 获取凸包顶点
hull_vertices = hull.vertices
# 提取凸包顶点坐标  
hull_points = uv_points[hull_vertices]

# 每个hue，只保留一个uv坐标。
hull_angle_to_points = {}  # 使用字典来存储每个色度角对应的唯一点
for idx in hull.vertices:
    angle = hue_angle_list[idx]
    if angle not in hull_angle_to_points:
        hull_angle_to_points[angle] = uv_points[idx]  # 只保留第一个uv坐标点

#  生成新的外轮廓点集合
new_contour_points = list(hull_angle_to_points.values())  # 取出唯一的uv坐标点
new_contour_points = np.array(new_contour_points)
# 按照色度角从小到大排序
sorted_angles = sorted(hull_angle_to_points.keys())  # 排序色度角
sorted_contour_points = []
for angle in sorted_angles:
    sorted_contour_points.append(hull_angle_to_points[angle])  # 按照排序的色度角收集对应的轮廓点
sorted_contour_points = np.array(sorted_contour_points)  # 转换为 NumPy 数组
# 添加色度角为0的uv点，确保轮廓线闭合
if sorted_contour_points.size > 0:
    first_point = sorted_contour_points[0]  # 获取第一个点
    sorted_contour_points = np.vstack([sorted_contour_points, first_point])  # 在末尾添加第一个点
    sorted_angles.append(sorted_angles[0])   # 末尾添加第一个点的hue值。

u_prime_with_hull = list(sorted_contour_points[:, 0])
v_prime_with_hull = list(sorted_contour_points[:, 1])

from openpyxl import load_workbook
output_filename = 'uv_gamut_1.xlsx'  # *************************************
sheet_name = '12640_LCH'


def write_excel(filename, sheet_name, hue_angle, uv_points):
    # 加载数据
    u_points = uv_points[:, 0]
    v_points = uv_points[:, 1]
    df_sorted_final = pd.DataFrame({
        'Hue Angle': hue_angle,  # 排序后的色度角
        'u\'': u_points,          # 排序后的 u' 坐标
        'v\'': v_points           # 排序后的 v' 坐标
    })

    # 加载 Excel 文件
    book = load_workbook(filename)
    
    # 删除同名 sheet（如果存在）
    if sheet_name in book.sheetnames:
        del book[sheet_name]

    # 确保至少有一个工作表是可见的
    if not any(sheet.sheet_state == 'visible' for sheet in book.worksheets):
        book.active.sheet_state = 'visible' # type: ignore

    with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
        writer._book = book  # type: ignore # 使用 _book 直接加载 openpyxl 的工作簿
        df_sorted_final.to_excel(writer, sheet_name=sheet_name, index=False)
        
uv_points = sorted_contour_points
write_excel(output_filename, sheet_name, sorted_angles, uv_points)        

"""
# 打开现有的Excel文件
book = load_workbook(output_filename)

# 如果存在同名的sheet，先删除
if sheet_name in book.sheetnames:
    del book[sheet_name]

# Create a DataFrame to store the sorted results
df_sorted_final = pd.DataFrame({
    'Hue Angle': sorted_angles,  # 排序后的色度角
    'u\'': u_prime_with_hull,          # 排序后的 u' 坐标
    'v\'': v_prime_with_hull           # 排序后的 v' 坐标
})

# 使用 pandas 和 openpyxl 写入排序后的数据
with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
    writer.book = book  # 载入已经存在的 workbook
    df_sorted_final.to_excel(writer, sheet_name=sheet_name, index=False)
"""

# 绘制 CIE 1976UCS 色域图
plot_chromaticity_diagram_CIE1976UCS(show_diagram_colours=False, standalone=False, color = 'orange') 

# 获取当前的坐标系 (ax)
ax = plt.gca()
plt.axis((-0.1, 0.7, -0.1, 0.7))
# Plot No Convex Hull curve
plt.plot(u_prime_no_hull, v_prime_no_hull, color = 'green', lw=2, label='No Convex Hull')    # 色度角hue下最大色饱和度Chroma 构成的轮廓。
plt.show()

plt.plot(u_prime_with_hull, v_prime_with_hull, color='orange', label='Concave Hull')   # 绘制凸包线条，每个hue只有一点在凸包线上。
plt.scatter(u_prime_list, v_prime_list, color='lightgray', s=15)  # 给出所有uv点图。 s 设置点的大小

# 标注色度角 (hue angle)， 当两标注点距离过小，则丢弃一个标注，避免两标注重叠。
min_distance = 0.02  # 可根据图的大小调整该值
# 记录上一个标注点的位置
last_annotated_point = None
# 标注色度角
for idx in range(len(sorted_contour_points)):
    hue_angle = sorted_angles[idx]
    u_point = sorted_contour_points[idx, 0]
    v_point = sorted_contour_points[idx, 1]
    # 检查两点距离，决定是否需要标注
    if last_annotated_point is None or np.linalg.norm(np.array([u_point, v_point]) - np.array(last_annotated_point)) > min_distance:
        # 在图中标注色度角（蓝色标注）
        plt.annotate(f'{hue_angle}°', (u_point, v_point), textcoords="offset points", xytext=(5, 5), color='blue', ha='center', fontsize=10)
        # 更新上一个标注点
        last_annotated_point = [u_point, v_point]

plt.title("pointer Gamut Boundary in u'v' Plane ")
plt.xlabel("u'")
plt.ylabel("v'")
plt.grid(True)
plt.legend()
plt.show()

print(f"Convex hull points saved to {output_filename}")