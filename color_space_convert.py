# 更新的色域图绘制，将D50，ILLUM_C颜色点变换成D65颜色点。chromatic_adaptation采用VonKries-Bradford方法, 改进的Convex Hull算法。
# python 3.12.3 测试通过，2024-12-19. matplotlib 后端默认开启QtAgg, 但找不到 Qt platform plugin "wayland" ，绘图黑，失效。
# 改为TkAgg后 matplotlib绘图正常。

import numpy as np
import colour
import pandas as pd
import matplotlib.pyplot as plt
from scipy.spatial import ConvexHull
from openpyxl import Workbook
from colour.plotting import plot_chromaticity_diagram_CIE1976UCS
from gamut_util import ConvexHull_general, ConvexHull_uni , LCHab_to_XYZ, LCHuv_to_XYZ, XYZ_to_uv, write_excel
import matplotlib

matplotlib.use('TkAgg')
plt.ion()
print("Currently Using Backend:", matplotlib.get_backend())  # 当前默认后端

ILLUM_C = colour.CCS_ILLUMINANTS["CIE 1931 2 Degree Standard Observer"]['C']  # CIE xy chromaticity coordinates
D65 = colour.CCS_ILLUMINANTS["CIE 1931 2 Degree Standard Observer"]["D65"]    # CIE xy chromaticity coordinates
D50 = colour.CCS_ILLUMINANTS["CIE 1931 2 Degree Standard Observer"]["D50"]    # CIE xy chromaticity coordinates

#from colour.adaptation import chromatic_adaptation
from colour.adaptation import chromatic_adaptation_VonKries

# 转换为XYZ白点 (假设 Y = 1)
def xy_to_XYZ(xy, Y=1.0):
    x, y = xy
    X = (Y / y) * x
    Z = (Y / y) * (1 - x - y)
    return np.array([X, Y, Z])

D50_white = xy_to_XYZ(D50)  # D50白点的XYZ值
D65_white = xy_to_XYZ(D65)  # D65白点的XYZ值
ILLUM_C_white = xy_to_XYZ(ILLUM_C)

from openpyxl import load_workbook
def delete_sheet_if_exists(file_path, sheet_name):
    # 使用 openpyxl 加载 Excel 文件
    workbook = load_workbook(file_path)
    # 如果目标工作表已经存在，则删除
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    # 保存删除后的文件
    workbook.save(file_path)

import colorspacious as cs
# 定义 CIELAB 色彩的映射函数
def hue_to_lab_color(hue_angle):
    # CIELAB 颜色空间的 a 和 b 轴可以用来表示色相
    L = 70  # 这里选择固定的亮度 L*
    C = 50  # 固定的色度 C*
    # 将 hue_angle 转换为弧度制并计算 a* 和 b* 值
    a = C * np.cos(np.radians(hue_angle))
    b = C * np.sin(np.radians(hue_angle))
    # CIELAB 颜色空间下的 (L*, a*, b*) 转换为 RGB
    lab_color = cs.cspace_convert([L, a, b], "CIELab", "sRGB1")
    # 确保 RGB 值在 [0, 1] 范围内
    lab_color = np.clip(lab_color, 0, 1)
    return lab_color

ILLUM_S = D50  #   确定标准光源  两坐标:  ILLUM_C,  D50,  D65,  ******************************************************
ILLUM_S_white = D50_white  #  三坐标:    ILLUM_C_white,  D50_white,  D65_white
file_path = 'LCH_gamut.xlsx'  #  Excel file   *****************************
sheet_name = '12640_LCHab'                           #  12640_LCHab,  TC300_LCHab,  pointer_LCHab, pointer_LCHuv    HP_printer,  PhotoGamut   *******************
source_format = 'LCHab'                            # LCHab,  LCHuv
# Read Excel data
df = pd.read_excel(file_path, sheet_name=sheet_name, index_col=0)
# Extract hue angles, lightness values, chroma values
hue = df.index.to_numpy()  # Hue angles (0-350, first column)
L_values = df.columns.to_numpy(dtype=float)  # Lightness values (5, 10, ..., 95, first row)
Chroma = df.to_numpy()  # Chroma values (36x19 array)

# hue: (36,) 数组，表示色相角
# L_values: (19,) 数组，表示亮度值
# Chroma: (36, 19) 数组，表示色相角与亮度对应的色度值
# 创建一个列表来存储结果
LCH_list = []
hue_angle_list = []
# 遍历每个色相角和亮度值的组合
for i, L in enumerate(L_values):  # 遍历每个亮度值
    for j, h in enumerate(hue):  # 遍历每个色相角
        C = Chroma[j, i]  # 对应的色度值 (Chroma value at (hue, L))
        # 创建 L, C, h 的数组
        LCH = np.array([L, C, h])
        # 将 LCH 数组添加到列表中
        LCH_list.append(LCH)
        hue_angle_list.append(h)
# 转换为 NumPy 数组，形状为 (36*19, 3)
LCH_array = np.array(LCH_list)
hue_angle_array = np.array(hue_angle_list)


# 1. No Convex Hull: Find the max C*ab for each hue angle and calculate u'v'
u_prime_no_hull = []
v_prime_no_hull = []
hue_angle_no_hull = []
for j, h in enumerate(hue):
    # Find the maximum C*ab for each hue angle across all lightness values
    max_C_ab = np.max(Chroma[j, :])
    L_max_index = np.argmax(Chroma[j, :])
    L_max = L_values[L_max_index]  # Get corresponding L* value
    LCH = np.array([L_max, max_C_ab, h])
    # 转换LCHab
    if source_format == 'LCHab':     # LCHab or LCHuv
        XYZ_data = LCHab_to_XYZ(LCH, ILLUM_S )  
    else:
        XYZ_data = LCHuv_to_XYZ(LCH, ILLUM_S )
    # Bradfor色度值转换，XYZ_D50 到 XYZ_D65
    XYZ_D65 = chromatic_adaptation_VonKries(XYZ_data, ILLUM_S_white, D65_white, transform="Bradford")  #  ***************************
    #XYZ_D65 = XYZ_data # ****************************************
    u_prime, v_prime = XYZ_to_uv(XYZ_D65)  
    # Store results
    u_prime_no_hull.append(u_prime)
    v_prime_no_hull.append(v_prime)
    hue_angle_no_hull.append(h)
# Now ensure the curve is closed by duplicating the 0-degree point at 360 degrees
u_prime_no_hull.append(u_prime_no_hull[0])
v_prime_no_hull.append(v_prime_no_hull[0])
hue_angle_no_hull.append(360)

#2  应用 ConvexHull
# 转换LCHab
if source_format == 'LCHab':   # LCHab or LCHuv
    XYZ_data = LCHab_to_XYZ(LCH_array, ILLUM_S )   
else:
    XYZ_data = LCHuv_to_XYZ(LCH_array, ILLUM_S )
# Bradfor色度值转换，XYZ_D50 到 XYZ_D65
XYZ_D65 = chromatic_adaptation_VonKries(XYZ_data, ILLUM_S_white, D65_white, transform="Bradford")  #  *************************   XYZ_D65
#XYZ_D65 = XYZ_data #  不采用chromatic_adaptation  变换**************************************************************
u_prime, v_prime = XYZ_to_uv(XYZ_D65)
uv_points = np.column_stack((u_prime, v_prime))

# 采用逐点逼近的凸包算法 convex_hull_general()
hull_angles, hull_points, closest_points= ConvexHull_general(hue_angle_list, uv_points)  # u_prime_list, v_prime_list

# 使用ConvexHull函数
hull = ConvexHull(uv_points)
# 获取凸包顶点
hull_vertices = hull.vertices
#提取凸包顶点坐标  
hull_points_1 = uv_points[hull_vertices]

# 绘制 CIE 1976UCS 色域图
plot_chromaticity_diagram_CIE1976UCS(show_diagram_colours=False, color ='lightgray', standalone=False) 
# 强制保持坐标轴比例为1:1
plt.gca().set_aspect('equal', adjustable='box')
# 获取当前的坐标系 (ax)
ax = plt.gca()
plt.axis((-0.1, 0.7, -0.1, 0.7))
plt.show()

# 绘制最大色调的轮廓
plt.plot(u_prime_no_hull, v_prime_no_hull, color='green', label='Max_Chromatic/hue')   
# 绘制convexhull外轮廓线
plt.plot(hull_points[:,0], hull_points[:,1], color='red', label='Convex Hull Modified')   # 绘制凸包线条
plt.plot(hull_points_1[:,0], hull_points_1[:,1], color='blue', label='Convex Hull', linestyle = '--')   # 绘制凸包线条


# 标注色度角 (hue angle)， 当两标注点距离过小，则丢弃一个标注，避免两标注重叠。
min_distance = 0.02  # 可根据图的大小调整该值
# 记录上一个标注点的位置
last_annotated_point = None
# 标注色度角
for idx in range(len(hull_points)):
    hue_angle = hull_angles[idx]
    u_point = hull_points[idx, 0]
    v_point = hull_points[idx, 1]
    # 检查两点距离，决定是否需要标注
    if last_annotated_point is None or np.linalg.norm(np.array([u_point, v_point]) - np.array(last_annotated_point)) > min_distance:
        # 在图中标注色度角（蓝色标注）
        plt.annotate(f'{hue_angle}°', (u_point, v_point), textcoords="offset points", xytext=(5, 5), color='blue', ha='center', fontsize=10)
        # 更新上一个标注点
        last_annotated_point = [u_point, v_point]

from matplotlib.colors import hsv_to_rgb
# 根据 hue_angle_list 生成 CIELAB 对应的颜色
colors = [hue_to_lab_color(hue) for hue in hue_angle_list]
# 在 u'v' 平面上绘制散点图
plt.scatter(u_prime, v_prime, color=colors, s=10)  # s 设置点的大小g

plt.title("CIE 1976UCS Chromaticity - CIE 1931 2° Standard Observer - Pointer LCHab")
plt.legend()

#filename = 'uv_gamut.xlsx'  # 变换结果(u, v)写入Excel
#sheet_name = 'HP_printer' 
#write_excel(filename, sheet_name, hue_angle, uv_points)

aa=1



