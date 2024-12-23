
#  CIE 1976UCS gamut 色域图公用函数
import numpy as np
from scipy.spatial import ConvexHull
from collections import defaultdict
import alphashape
from shapely.geometry import Polygon
import colour
import pandas as pd
import matplotlib.pyplot as plt

# 读取 Excel 数据
def read_excel(filename, sheet_name):
    df = pd.read_excel(io=filename, sheet_name=sheet_name, index_col=0)

    # 提取色调角度、明度值、和饱和度值
    hue = df.index.to_numpy()  # 色调角度 (h_uv, 第一列)
    L_values = df.columns.to_numpy(dtype=float)  # 明度值 (L*, 第一行)
    Chroma = df.to_numpy()  # 饱和度值 (C_uv, 36x16 数组, 576点)

    # hue: (36,) 数组，表示色相角
    # L_values: (19,) 数组，表示亮度值
    # Chroma: (36, 19) 数组，表示色相角与亮度对应的色度值
    LCH_list = []
    hue_angle_list = []
    # 遍历每个色相角和亮度值的组合
    for i, L in enumerate(L_values):  # 遍历每个亮度值
        for j, h in enumerate(hue):  # 遍历每个色相角
            C = Chroma[j, i]  # 对应的色度值 (Chroma value at (hue, L))
            # 创建 L, C, h 的数组
            LCH = np.array([L, C, h])
            # 将 LCH 数组添加到列表中
            LCH_list.append(LCH)
            hue_angle_list.append(h)
    # 转换为 NumPy 数组，形状为 (36*19, 3)
    LCH_array = np.array(LCH_list)
    hue_angle_array = np.array(hue_angle_list)
    return L_values,  LCH_array, hue_angle_array


from openpyxl import load_workbook
def write_excel(filename, sheet_name, hue_angle, uv_points):
    # 加载数据
    u_points = uv_points[:, 0]
    v_points = uv_points[:, 1]
    df_sorted_final = pd.DataFrame({
        'Hue Angle': hue_angle,  # 排序后的色度角
        'u\'': u_points,          # 排序后的 u' 坐标
        'v\'': v_points           # 排序后的 v' 坐标
    })

    # 加载 Excel 文件
    book = load_workbook(filename)
    
    # 删除同名 sheet（如果存在）
    if sheet_name in book.sheetnames:
        del book[sheet_name]

    # 确保至少有一个工作表是可见的
    if not any(sheet.sheet_state == 'visible' for sheet in book.worksheets):
        book.active.sheet_state = 'visible' # type: ignore

    with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
        writer._book = book  # type: ignore # 使用 _book 直接加载 openpyxl 的工作簿
        df_sorted_final.to_excel(writer, sheet_name=sheet_name, index=False)


# 标注色度角 (hue angle)， 当两标注点距离过小，则丢弃一个标注，避免两标注重叠。
def annotate_hue_angle(hue_angle, uv_points):
    min_distance = 0.02  # 可根据图的大小调整该值
    # 记录上一个标注点的位置
    last_annotated_point = None
    # 标注色度角
    for idx in range(len(uv_points)):
        hue_angle_tmp = hue_angle[idx]
        u_point = uv_points[idx, 0]
        v_point = uv_points[idx, 1]
        # 检查两点距离，决定是否需要标注
        if last_annotated_point is None or np.linalg.norm(np.array([u_point, v_point]) - np.array(last_annotated_point)) > min_distance:
            # 在图中标注色度角（蓝色标注）
            plt.annotate(f'{hue_angle_tmp}°', (u_point, v_point), textcoords="offset points", xytext=(5, 5), color='blue', ha='center', fontsize=10)
            # 更新上一个标注点
            last_annotated_point = [u_point, v_point]

# convex hull算法的第一种变化
def ConvexHull_general(hue_angle_list, uv_points ):

    # 1. 计算凸包
    #uv_points = np.column_stack((u_prime_list, v_prime_list))
    u_prime_list =uv_points[:,0]
    v_prime_list = uv_points[:,1]
    hull = ConvexHull(uv_points)
    hull_points = uv_points[hull.vertices]  # 获取凸包上的点

    # 2. 找出被忽略的色度角
    hull_angles = []
    for point_temp in hull_points:
        idx = np.where((point_temp == uv_points).all(axis=1))[0][0]
        hull_angles.append(hue_angle_list[idx])

    # 找出所有的色度角，排除凸包上的色度角
    unique_hue_angles = set(hue_angle_list)  # 提取唯一色度角
    ignored_angles = list(unique_hue_angles.difference(hull_angles))

    # 创建角度与点的映射
    angle_to_points = defaultdict(list)
    for i, angle in enumerate(hue_angle_list):
        angle_to_points[angle].append((u_prime_list[i], v_prime_list[i]))

    # 3. 计算忽略点到凸包的距离

    def point_to_line_dist(point, line_start, line_end):
        """计算点 point 到线段 (line_start, line_end) 的最短距离"""
        line_vec = line_end - line_start
        point_vec = point - line_start
        line_len = np.linalg.norm(line_vec)
        if line_len == 0:
            return np.linalg.norm(point_vec)
        proj_len = np.dot(point_vec, line_vec) / line_len
        proj_len = np.clip(proj_len, 0, line_len)
        proj_point = line_start + (proj_len / line_len) * line_vec
        return np.linalg.norm(point - proj_point)

    # 对每个忽略的色度角，找最接近的外轮廓点
    closest_points = []
    for angle in ignored_angles:
        angle_points = angle_to_points[angle]  # 获取该色度角的所有点
        min_dist = float('inf')  # 初始化最小距离
        closest_point = None  # 初始化最近点
        # 对于该色度角的所有点，找到距离外轮廓最近的那个点
        for ignored_point in angle_points:
            for i in range(len(hull.vertices)):
                start = hull_points[i]
                end = hull_points[(i + 1) % len(hull.vertices)]
                # 计算该忽略点到外轮廓线段的距离
                dist = point_to_line_dist(np.array(ignored_point), start, end)
                # 找到距离最小的点
                if dist < min_dist:
                    min_dist = dist
                    closest_point = ignored_point  # 更新为当前最近的点
        if closest_point is not None:
            closest_points.append((angle, closest_point))  # 每个色度角只保存一个最近点

    # 4. 将凸包点与色度角对应
    hull_angle_to_points = {}  # 使用字典来存储每个色度角对应的唯一点
    for idx in hull.vertices:
        angle = hue_angle_list[idx]
        if angle not in hull_angle_to_points:
            hull_angle_to_points[angle] = uv_points[idx]  # 只保留第一个uv坐标点

    # 5. 添加最接近的忽略点，确保每个色度角都有一个对应的外轮廓点
    for angle, closest_point in closest_points:
        if angle not in hull_angle_to_points:
            hull_angle_to_points[angle] = closest_point  # 添加忽略的点

    # 6. 生成新的外轮廓点集合
    new_contour_points = list(hull_angle_to_points.values())  # 取出唯一的uv坐标点

    # 7. 按照色度角从小到大排序
    sorted_angles = sorted(hull_angle_to_points.keys())  # 排序色度角
    sorted_contour_points = []
    for angle in sorted_angles:
        sorted_contour_points.append(hull_angle_to_points[angle])  # 按照排序的色度角收集对应的轮廓点

    sorted_contour_points = np.array(sorted_contour_points)  # 转换为 NumPy 数组

    # 8. 添加色度角为0的uv点，确保轮廓线闭合
    if sorted_contour_points.size > 0:
        first_point = sorted_contour_points[0]  # 获取第一个点
        first_point_hue = sorted_angles[0]
        sorted_contour_points = np.vstack([sorted_contour_points, first_point])  # 在末尾添加第一个点
        sorted_angles = np.append(sorted_angles, first_point_hue)
    return sorted_angles, sorted_contour_points, closest_points  # 返回排序后的结果
#  *******************************************************************************************************************

# 常规 ConvexHull函数第二种改变,   input uv_points,  no hue_angle_array 
def ConvexHull_uni(hue_angle_array, uv_points):
    # 使用ConvexHull函数
    hull = ConvexHull(uv_points)
    # 获取凸包顶点
    hull_vertices = hull.vertices
    # 提取凸包顶点坐标  
    hull_points = uv_points[hull_vertices]

    # 从 hull_vertices 找到对应的 hue 角度
    hull_hue_angles = hue_angle_array[hull_vertices]  # 对应的 hue 角度
   # 获取排序的索引
    sorted_indices = np.argsort(hull_hue_angles)
    # 根据排序索引对 hull_hue_angles 和 hull_points 进行排序
    sorted_hull_hue_angles = hull_hue_angles[sorted_indices]
    sorted_hull_points = hull_points[sorted_indices]

        # 获取去重后的 hue 角度及其索引
    unique_hull_hue_angles, unique_indices = np.unique(sorted_hull_hue_angles, return_index=True)
    # 根据 unique_indices 对 sorted_hull_points 进行去重
    unique_hull_points = sorted_hull_points[unique_indices]

    # 添加色度角为0的uv点, 确保轮廓线闭合
    if unique_hull_points.size > 0:
        first_point = unique_hull_points[0]  # 获取第一个点
        unique_hull_points = np.vstack([unique_hull_points, first_point])  # 在末尾添加第一个点
        # 使用 np.append 在末尾添加第一个点的 hue 值
        unique_hull_hue_angles = np.append(unique_hull_hue_angles, unique_hull_hue_angles[0])   # 
    return unique_hull_hue_angles, unique_hull_points

def XYZ_to_uv(XYZ):
    # 确保 XYZ 至少是二维数组，形状变为 (n, 3)
    XYZ = np.atleast_2d(XYZ)# 提取 X, Y, Z 列
    X = XYZ[:,0]
    Y = XYZ[:,1]
    Z = XYZ[:,2]
    # 计算分母：X + 15Y + 3Z
    denominator = X + 15 * Y + 3 * Z
    # 避免除以零的情况，添加一个非常小的值 epsilon
    epsilon = 1e-10
    denominator = np.maximum(denominator, epsilon)
    # 计算 u' 和 v' 按公式
    u_prime = 4 * X / denominator
    v_prime = 9 * Y / denominator
    return u_prime, v_prime

def LCHab_to_XYZ(LCHab, ILLUMINATS):
    Lab = colour.LCHab_to_Lab(LCHab)
    XYZ = colour.Lab_to_XYZ(Lab, ILLUMINATS)  
    return XYZ

def LCHab_to_uv(LCHab, ILLUMINATS):
    Lab = colour.LCHab_to_Lab(LCHab)
    XYZ = colour.Lab_to_XYZ(Lab, ILLUMINATS)  
    return XYZ_to_uv(XYZ)

def LCHuv_to_XYZ(LCHuv, ILLUMINATS):
    Luv = colour.LCHuv_to_Luv(LCHuv)
    XYZ = colour.Luv_to_XYZ(Luv, ILLUMINATS)  
    return XYZ

def Lab_to_LCHab(Lab):
    # 确保 Lab 是二维数组，形状为 (n, 3)
    Lab = np.atleast_2d(Lab)
     # 提取 L, a, b 列
    L = Lab[:, 0]
    a = Lab[:, 1]
    b = Lab[:, 2]
    # 计算 Chroma (C*)
    C = np.sqrt(a**2 + b**2)
    # 计算 Hue Angle (h_ab) 并将其转换为角度制
    h = np.degrees(np.arctan2(b, a))
    # 保证 hue 角度在 0 到 360 之间
    h[h < 0] += 360
    # 组合 L, C*, h_ab 结果为 (n, 3) 数组
    LCHab = np.stack([L, C, h], axis=1)
    return LCHab

def Alpha_outline(hue_angle_list, u_list, v_list, alpha_value):
    # 第一段：生成 Alpha Shape 并提取外轮廓点
    #alpha_value = 10.0
    uv_points = np.column_stack((u_list, v_list))
    alpha_shape = alphashape.alphashape(uv_points, alpha_value)

    # 提取外轮廓的 x, y 坐标
    if isinstance(alpha_shape, Polygon):
        x, y = alpha_shape.exterior.xy

    # 将 u_list 和 v_list 转换为数组
    u_list = np.array(u_list)
    v_list = np.array(v_list)

    # 第二段：匹配每个外轮廓点 x, y，找到对应的 hue_angle，并排除重复 hue_angle
    matched_hue_angles = []
    unique_hue_points = []

    # 遍历每个外轮廓点 x, y
    for i in range(len(x)):
        # 查找满足条件的索引
        indices = np.where((u_list == x[i]) & (v_list == y[i]))[0]

        if indices.size > 0:  # 确保找到了匹配的点
            index = indices[0]  # 取第一个匹配的索引
            matched_hue_angle = hue_angle_list[index]

            # 检查 hue_angle 是否已经处理过，避免重复
            if matched_hue_angle not in matched_hue_angles:
                matched_hue_angles.append(matched_hue_angle)
                unique_hue_points.append((matched_hue_angle, x[i], y[i]))
    
    return np.array(unique_hue_points)

